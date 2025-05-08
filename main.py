from flask import Flask, request, send_file, render_template_string
import openpyxl
import io

app = Flask(__name__)

@app.route("/", methods=["GET"])
def index():
    return render_template_string(open("templates/index.html").read())

@app.route("/guardar", methods=["POST"])
def guardar():
    dato1 = request.form.get("dato1", type=float)
    dato2 = request.form.get("dato2", type=float)
    dato3 = request.form.get("dato3", type=float)

    wb = openpyxl.load_workbook("Prueba.xlsx")
    ws = wb.active

    ws["B3"] = dato1
    ws["B4"] = dato2
    ws["B5"] = dato3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="Datos_Completados.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
